import io

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

from app.core.time import now_pkt
from app.schemas.dashboard_export import DashboardExportRequest

SECTION_FILL = PatternFill(start_color="EEF2FF", end_color="EEF2FF", fill_type="solid")
HEADER_FONT = Font(bold=True, size=11)
TITLE_FONT = Font(bold=True, size=16)
SUBTITLE_FONT = Font(size=10, color="666666")


def _section(ws, row: int, title: str) -> int:
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = HEADER_FONT
    cell.fill = SECTION_FILL
    ws.cell(row=row, column=2).fill = SECTION_FILL
    return row + 1


def _kv(ws, row: int, label: str, value: str) -> int:
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=2, value=value)
    return row + 1


def build_excel(data: DashboardExportRequest) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 16

    row = 1
    title_cell = ws.cell(row=row, column=1, value="ORBIT — Company Overview")
    title_cell.font = TITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Period: {data.period_label}  ·  Currency: {data.currency_label}  ·  {data.fx_note}").font = SUBTITLE_FONT
    row += 1
    ws.cell(row=row, column=1, value=f"Generated {now_pkt().strftime('%d %b %Y, %I:%M %p')} PKT").font = SUBTITLE_FONT
    row += 2

    row = _section(ws, row, "Revenue")
    row = _kv(ws, row, "Won & contracted (locked)", data.revenue.locked)
    row = _kv(ws, row, "Invoiced, not yet collected", data.revenue.invoiced)
    row = _kv(ws, row, "Collected", data.revenue.collected)
    row = _kv(ws, row, "Expected revenue (pipeline, stage-weighted)", data.revenue.expected)
    row += 1

    row = _section(ws, row, "Cash Position")
    row = _kv(ws, row, "Owed to us (receivables)", data.cash_position.receivables)
    row = _kv(ws, row, "Payroll / month", data.cash_position.payroll_month)
    row = _kv(ws, row, "Total cash-out / month", data.cash_position.total_cash_out_month)
    row = _kv(ws, row, "Net position (collected − out)", data.cash_position.net_position)
    row = _kv(ws, row, "Expenses this month", data.expenses_month)
    row += 1

    if data.delayed_projects:
        row = _section(ws, row, "Delayed Projects")
        ws.cell(row=row, column=1, value="Project").font = HEADER_FONT
        ws.cell(row=row, column=2, value="Client").font = HEADER_FONT
        ws.cell(row=row, column=3, value="Days overdue").font = HEADER_FONT
        row += 1
        for p in data.delayed_projects:
            ws.cell(row=row, column=1, value=p.name)
            ws.cell(row=row, column=2, value=p.client or "")
            ws.cell(row=row, column=3, value=p.days_overdue or "")
            row += 1
        row += 1

    if data.profitability:
        row = _section(ws, row, "Project Profitability")
        ws.cell(row=row, column=1, value="Project").font = HEADER_FONT
        ws.cell(row=row, column=2, value="Revenue").font = HEADER_FONT
        ws.cell(row=row, column=3, value="Cost").font = HEADER_FONT
        ws.cell(row=row, column=4, value="Margin").font = HEADER_FONT
        row += 1
        for p in data.profitability:
            ws.cell(row=row, column=1, value=p.name)
            ws.cell(row=row, column=2, value=p.revenue)
            ws.cell(row=row, column=3, value=p.cost)
            ws.cell(row=row, column=4, value=p.margin)
            row += 1
        row += 1

    if data.utilization:
        row = _section(ws, row, "Resource Utilization")
        ws.cell(row=row, column=1, value="Employee").font = HEADER_FONT
        ws.cell(row=row, column=2, value="Allocation").font = HEADER_FONT
        row += 1
        for u in data.utilization:
            ws.cell(row=row, column=1, value=u.name)
            ws.cell(row=row, column=2, value=u.pct)
            row += 1
        row += 1

    if data.category_budgets:
        row = _section(ws, row, "Expenses by Category — Budget vs. Actual")
        ws.cell(row=row, column=1, value="Category").font = HEADER_FONT
        ws.cell(row=row, column=2, value="Actual").font = HEADER_FONT
        ws.cell(row=row, column=3, value="Budget").font = HEADER_FONT
        row += 1
        for b in data.category_budgets:
            ws.cell(row=row, column=1, value=b.category)
            ws.cell(row=row, column=2, value=b.actual)
            ws.cell(row=row, column=3, value=b.budget)
            row += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(data: DashboardExportRequest) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=letter,
        topMargin=0.6 * inch, bottomMargin=0.6 * inch,
        leftMargin=0.6 * inch, rightMargin=0.6 * inch,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("ORBITTitle", parent=styles["Title"], fontSize=20, spaceAfter=4)
    subtitle_style = ParagraphStyle("ORBITSubtitle", parent=styles["Normal"], textColor=colors.HexColor("#666666"), spaceAfter=14)
    section_style = ParagraphStyle("ORBITSection", parent=styles["Heading2"], fontSize=13, spaceBefore=16, spaceAfter=8, textColor=colors.HexColor("#4F46E5"))

    story = [
        Paragraph("ORBIT — Company Overview", title_style),
        Paragraph(
            f"Period: {data.period_label} &nbsp;&middot;&nbsp; Currency: {data.currency_label} &nbsp;&middot;&nbsp; {data.fx_note}"
            f"<br/>Generated {now_pkt().strftime('%d %b %Y, %I:%M %p')} PKT",
            subtitle_style,
        ),
    ]

    def kv_table(rows):
        t = Table(rows, colWidths=[3.2 * inch, 3.4 * inch])
        t.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("FONTNAME", (1, 0), (1, -1), "Helvetica-Bold"),
        ]))
        return t

    def data_table(header, rows, col_widths):
        table_data = [header] + rows
        t = Table(table_data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EEF2FF")),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ]))
        return t

    story.append(Paragraph("Revenue", section_style))
    story.append(kv_table([
        ["Won & contracted (locked)", data.revenue.locked],
        ["Invoiced, not yet collected", data.revenue.invoiced],
        ["Collected", data.revenue.collected],
        ["Expected revenue (pipeline, stage-weighted)", data.revenue.expected],
    ]))

    story.append(Paragraph("Cash Position", section_style))
    story.append(kv_table([
        ["Owed to us (receivables)", data.cash_position.receivables],
        ["Payroll / month", data.cash_position.payroll_month],
        ["Total cash-out / month", data.cash_position.total_cash_out_month],
        ["Net position (collected − out)", data.cash_position.net_position],
        ["Expenses this month", data.expenses_month],
    ]))

    if data.delayed_projects:
        story.append(Paragraph("Delayed Projects", section_style))
        story.append(data_table(
            ["Project", "Client", "Days overdue"],
            [[p.name, p.client or "—", p.days_overdue or "—"] for p in data.delayed_projects],
            [2.6 * inch, 2.2 * inch, 1.8 * inch],
        ))

    if data.profitability:
        story.append(Paragraph("Project Profitability", section_style))
        story.append(data_table(
            ["Project", "Revenue", "Cost", "Margin"],
            [[p.name, p.revenue, p.cost, p.margin] for p in data.profitability],
            [2.6 * inch, 1.4 * inch, 1.4 * inch, 1.2 * inch],
        ))

    if data.utilization:
        story.append(Paragraph("Resource Utilization", section_style))
        story.append(data_table(
            ["Employee", "Allocation"],
            [[u.name, u.pct] for u in data.utilization],
            [4 * inch, 2.6 * inch],
        ))

    if data.category_budgets:
        story.append(Paragraph("Expenses by Category — Budget vs. Actual", section_style))
        story.append(data_table(
            ["Category", "Actual", "Budget"],
            [[b.category, b.actual, b.budget] for b in data.category_budgets],
            [2.6 * inch, 2.2 * inch, 2.2 * inch],
        ))

    story.append(Spacer(1, 4))
    doc.build(story)
    return buf.getvalue()
